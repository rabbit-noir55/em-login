#!/usr/bin/env python3
#pip install requests beautifulsoup4 easyocr numpy opencv-python pillow rich openpyxl
from rich.console import Console
from rich.prompt import Prompt
import json
import os
from datetime import datetime
from openpyxl import load_workbook
import  requests 
import json

import os

def get_data_from_json(keyword):
    """
    JSON fayl ichidan berilgan keyword bo'yicha qiymatni qaytaradi.
    Agar fayl mavjud bo'lmasa yoki notoâ€˜gâ€˜ri formatda boâ€˜lsa, yangi fayl yaratadi.
    
    :param keyword: Qidirilayotgan kalit soâ€˜z (str)
    :return: Kalit soâ€˜zga mos qiymat yoki None
    """
    file_path = os.path.join(os.path.dirname(__file__), "data.json")

    # JSON faylni tekshirish va yuklash
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
        except json.JSONDecodeError:
            print("[bold red]Xatolik: JSON fayl notoâ€˜gâ€˜ri formatda! Qayta yaratilyapti...[/bold red]")
            data = {"a": [], "b": []}
            with open(file_path, "w", encoding="utf-8") as file:
                json.dump(data, file, indent=4, ensure_ascii=False)
    else:
        print("[bold yellow]Ogohlantirish: data.json topilmadi! Yangi fayl yaratilmoqda...[/bold yellow]")
        data = {"a": [], "b": []}
        with open(file_path, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=4, ensure_ascii=False)

    return data.get(keyword, None)

# Funksiyani test qilish


import requests
import easyocr
import numpy as np
import cv2
from io import BytesIO
from PIL import Image

def preprocess_image(image):
    """Rasmni tozalash va kontrastni oshirish"""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # Oq-qora formatga o'tkazish
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)  # Shovqinni kamaytirish
    _, thresh = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)  # Kontrastni oshirish
    return thresh

def solve_captcha(image_url):
    """Yangi OCR modeli yordamida CAPTCHA ni aniqroq yechish"""
    response = requests.get(image_url)
    img = Image.open(BytesIO(response.content))
    img = np.array(img)

    # Rasmni oldindan ishlov berish
    processed_img = preprocess_image(img)

    # OCR yordamida raqamlarni o'qish
    reader = easyocr.Reader(['en'], gpu=False)
    result = reader.readtext(processed_img, detail=0, contrast_ths=0.8, adjust_contrast=0.7, decoder='beamsearch')

    # Faqat sonlarni ajratib olish
    captcha_text = ''.join(filter(str.isdigit, ''.join(result)))
    return captcha_text

import time
import base64
import numpy as np
from io import BytesIO
from PIL import Image
import easyocr
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


def faqat_raqamlar(matn: str) -> str:
    """Extracts numbers from the text."""
    return ''.join(filter(str.isdigit, matn))


def process_captcha(image):
    """Uses EasyOCR to read the CAPTCHA."""
    reader = easyocr.Reader(["en"])
    result = reader.readtext(np.array(image), detail=0)
    return faqat_raqamlar(result)


def auto_login(login, password):
    xato=[]
    """Automates login with retry attempts for each login/password combination."""
    max_attempts = 8
    for attempt in range(max_attempts):
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.get("https://login.emaktab.uz/")
        time.sleep(5)

        driver.find_element(By.NAME, "login").send_keys(login)
        driver.find_element(By.NAME, "password").send_keys(password)
        driver.find_element(By.NAME, "password").submit()

        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])
            time.sleep(3)

        try:
            captcha_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//img[@alt='captcha image']"))
            )

            captcha_text = ""
            for captcha_attempt in range(1):  # Retry 3 times for CAPTCHA if OCR fails
                captcha_url = captcha_element.get_attribute("src")

                if "data:image" in captcha_url:
                    captcha_base64 = captcha_url.split(",")[1]
                    captcha_bytes = base64.b64decode(captcha_base64)
                    captcha_img = Image.open(BytesIO(captcha_bytes))
                else:
                    captcha_element.screenshot("captcha.png")
                    captcha_img = Image.open("captcha.png")

                captcha_text = process_captcha(captcha_img)

                if captcha_text:
                    print(f"ğŸ” CAPTCHA topildi: {captcha_text}")
                    break
                else:
                    print(f"ğŸ”„ CAPTCHA topilmadi, {captcha_attempt + 1}-urinishda qayta urinib koâ€˜rilmoqda...")
                    time.sleep(5)

            driver.find_element(By.XPATH, "//input[@title='Rasmdagi belgilarni kiriting']").send_keys(captcha_text)
            driver.find_element(By.XPATH, "//input[@title='Rasmdagi belgilarni kiriting']").submit()

            time.sleep(5)

            current_url = driver.current_url
            if "userfeed" in current_url:
                print(f"âœ… {login} uchun muvaffaqiyatli login qildik!")
                driver.quit()
                break
            else:
                print(f"âŒ {login} uchun login amalga oshmadi! CAPTCHA yoki boshqa muammo boâ€˜lishi mumkin.")
                if attempt == max_attempts - 1:
                    print(f"â— {login} uchun barcha urinishlar tugadi. Keyingi login parolga o'tildi.")
                    xato.append(login)
                    driver.quit()
                    break
                driver.quit()
                time.sleep(5)

        except Exception as e:
            print(f"âŒ {login} uchun xatolik yuz berdi: {str(e)}")
            driver.quit()
            time.sleep(5)

    print(xato,"ushbu login xato qattiq urinildi lekin kirolmadik )
# Login and password lists
login_list = ["dimamatdinov"]

password_list = ["imamatdinov11"]

def mm():

    for login, password in zip(login_list, password_list):
        auto_login(login, password)

def x():
    
    # Login qilish uchun saytning toâ€˜gâ€˜ri URL manzilini kiriting
    url = "https://login.emaktab.uz/login"

    # Foydalanuvchi ma'lumotlari
    logins = get_data_from_json("a")
    passwords = get_data_from_json("b")

    # HTTP sarlavhalari
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Content-Type": "application/x-www-form-urlencoded"
    }
    result_login=[]
    icon=[]
    # Har bir foydalanuvchi uchun login qilish
    for login, password in zip(logins, passwords):
        session = requests.Session()
        
        # Saytga yuboriladigan ma'lumotlar
        data = {
            "login": login,
            "password": password
        }
        
        # POST soâ€˜rov yuborish
        response = session.post(url, data=data, headers=headers)

        
        # Natijani tekshirish
        if "userfeed" in response.url:
            result_login.append(login)
            icon.append("âœ…")
            
        else:
            result_login.append(login)
            icon.append("âŒ")
    return result_login, icon
console = Console()

def create_empty_data():
    return {"a": [], "b": []}

def load_data():
    if not os.path.exists("data.json"):
        return create_empty_data()
    
    with open("data.json", "r") as file:
        return json.load(file)

def save_data(data):
    with open("data.json", "w") as file:
        json.dump(data, file)

def check_program_expiry():
    expiry_date = datetime(2025, 5, 25)
    if datetime.now() > expiry_date:
        console.print("[bold red]Dastur muddati tugadi![/bold red]")
        exit()

def load_excel_data():
    filename = "data.xlsx"
    if not os.path.exists(filename):
        console.print("[bold red]Excel fayli topilmadi, iltimos faylni qo'shing![/bold red]")
        return
    
    wb = load_workbook(filename)
    ws = wb.active
    
    data = load_data()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            data["a"].append(str(row[0]))
            data["b"].append(str(row[1]))
    
    save_data(data)
    wb.close()
    os.remove(filename)
    console.print("[bold green]Excel ma'lumotlari yuklandi va fayl oâ€˜chirildi![/bold green]")

def show_menu():
    console.print("[bold green]Asosiy Menu[/bold green]")
    console.print("â•”â•â•â•â•¤â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•—")
    console.print("â•‘ K â”‚ Boâ€˜lim                                            â•‘")
    console.print("â•Ÿâ”€â”€â”€â”¼â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â•¢")
    console.print("â•‘ 1 â”‚ logon qilish  funksiyasini ishga tushirish        â•‘")
    console.print("â•‘ 2 â”‚ 90 logon qilish  funksiyasini ishga tushirish     â•‘")
    console.print("â•‘ 3 â”‚ Ma'lumotni tahrirlash                             â•‘")
    console.print("â•‘ 4 â”‚ Ma'lumotlarni oâ€˜chirish                           â•‘")
    console.print("â•‘ 5 â”‚ Excel'dan ma'lumot yuklash                        â•‘")
    console.print("â•‘ 6 â”‚ Chiqish                                           â•‘")
    console.print("â•šâ•â•â•â•§â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•")

def x_function():
    login_m,iconm_m=x()
    for log,icon in zip(login_m,iconm_m):
        print(log,icon)
    console.print("[bold cyan]logon qilish funksiyasi ishga tushdi![/bold cyan]")

def edit_data_section():
    data = load_data()
    if not data["a"]:
        console.print("[bold red]Ma'lumotlar yoâ€˜q![/bold red]")
        return

    for idx, (a, b) in enumerate(zip(data["a"], data["b"]), 1):
        console.print(f"{idx}. {a} | {b}")

    index = int(Prompt.ask("Tahrir qilmoqchi bo'lgan ma'lumotning K raqamini kiriting:"))
    if index < 1 or index > len(data["a"]):
        console.print("[bold red]Notoâ€˜gâ€˜ri K raqami![/bold red]")
        return

    new_a_value = Prompt.ask("Yangi A qiymatini kiriting:")
    new_b_value = Prompt.ask("Yangi B qiymatini kiriting:")
    data["a"][index - 1] = new_a_value
    data["b"][index - 1] = new_b_value
    save_data(data)
    console.print("[bold green]Ma'lumot muvaffaqiyatli tahrirlandi![/bold green]")

def delete_data_section():
    data = load_data()
    if not data["a"]:
        console.print("[bold red]Ma'lumotlar yoâ€˜q![/bold red]")
        return

    for idx, (a, b) in enumerate(zip(data["a"], data["b"]), 1):
        console.print(f"{idx}. {a} | {b}")
    
    index = int(Prompt.ask("Oâ€˜chirmoqchi boâ€˜lgan ma'lumotning K raqamini kiriting:"))
    if index < 1 or index > len(data["a"]):
        console.print("[bold red]Notoâ€˜gâ€˜ri K raqami![/bold red]")
        return
    
    data["a"].pop(index - 1)
    data["b"].pop(index - 1)
    save_data(data)
    console.print("[bold green]Ma'lumot muvaffaqiyatli oâ€˜chirildi![/bold green]")

def main_menu():
    check_program_expiry()
    while True:
        show_menu()
        choice = Prompt.ask("Tanlang:", choices=["1", "2", "3", "4", "5", "6"])
        if choice in "1":
            x_function()
        if choice in "2":
            mm()
        elif choice == "3":
            edit_data_section()
        elif choice == "4":
            delete_data_section()
        elif choice == "5":
            load_excel_data()
        elif choice == "6":
            console.print("[bold green]Dasturdan chiqildi![/bold green]")
            break
        else:
            console.print("[bold red]Notoâ€˜gâ€˜ri tanlov![/bold red]")

if __name__ == "__main__":
    main_menu()
